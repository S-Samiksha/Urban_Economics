import pandas as pd
import ast
#f = open("./busstopv2.txt", "r")
#pd.read_csv(pd.compat.StringIO("\.join(lines)), sep=";")
with open("busstopv2.txt", "r") as grilled_cheese:
    lines = grilled_cheese.readlines()
    i=0
    collector = []
    for item in lines:
        if i < 1000:
            collector.append(ast.literal_eval(item))
            #df = dict(item)
            #print(df)
            #print(type(df))
            i+=1
    print(len(collector))   
    #dictlines = dict(lines)
    #print(dictlines["value"])

i=2
df = pd.DataFrame(collector[i]['value'],index=[i for i in range(len(collector[1]['value']))])
print(df)
	
import openpyxl
dfs = []
for i in range(len(collector)):
    dfs.append(pd.DataFrame(collector[i]['value'],index=[i for i in range(len(collector[i]['value']))]))
print(f"Done with appending {i} DataFrames")
for i in range(len(collector)):
    #pass
    print(f"Dataframe {i+1}:\n")
    print(dfs[i])
    
overall = pd.concat(dfs,axis=0)
print("Overall DataFrame:")
print(overall.shape)

### Now that DataFrame is successfully created, output into Excel spreadsheet:"
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# For each row in DataFrame
for r in dataframe_to_rows(overall, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

#wb.save("busstop_overall.xlsx")

### Proceed to compare distance from HDB to each bus stop.
#Load all HDB data first:

allhdb = pd.read_excel("./hdb_to_mrt_all.xlsx", index_col=0,engine='openpyxl') #import fresh data  
global blocks
blocks = allhdb[['postal','lng_hdb','lat_hdb']].drop_duplicates()

# Find relevant long and lat for HDB (do not mix up with MRT)
print("Blocks DataFrame has shape:",blocks.shape)

#print(blocks)

#blocks = blocks[blocks['postal']=='650383']

def distance(long1,lat1,long2,lat2):
    # 1 deg = 111000m
    dist = 111000*((long2-long1)**2 +(lat2-lat1)**2)**0.5
    return dist

def all_distance(flist, overall):
    '''Returns closest bus stop from each HDB block in Singapore'''
    # For each block
    all_dist = {}
    min_dist = {}
    alldist_df = pd.DataFrame()

    margin_h = 0.002
    margin_v = 0.002  # of a degree. 1 degree = 111000m
    # 0.004 = 444m region
    box ={}
    # flight - index
    # fdata - block's long and lat
    for flight, fdata in flist.iterrows():
        #print(f"fdata is:\\n{fdata}")
        #print(f"{fdata.shape}")
        print(f"Working on {fdata['postal']}")
        long1 = fdata['lng_hdb']
        lat1 = fdata['lat_hdb']
        # Compare row-by-row with all bus stops:
        print(type(long1),type(lat1))
        box[flight]=[]
        for row, busstop in overall.iterrows():
            #print(busstop)
            long2 = busstop['Longitude']
            lat2 = busstop['Latitude']
            #print(f"Long2 is {long2}, Lat2 is {lat2}")
            #print(type(long2),type(lat2))
            if abs(long2-long1) > margin_h:
                continue
            elif abs(lat2-lat1) > margin_v:
                continue
            else:
                busstop['Distance']  = distance(long1,lat1,long2,lat2)
                busstop['Postal'] = fdata['postal']
                #print(f"Bus stop {busstop['Description']}: {busstop['Distance']}m")
                box[flight].append(pd.DataFrame(busstop).T)
                #index=[busstop['BusStopCode']
        
        print("Currently the list of bus stops are:",box[flight])
            #print(f"Distance is {dist}m")

            #dist = math.sqrt(ellipdist(long1, lat1, long, lat)**2)
            #rowdata.insert(loc=0, column='Distance (ft)', value=dist_ft)
        for key, value in box.items():
            #print("The value is",value)
            min_dist[key]= pd.concat(value,axis=0)
            min_dist[key].dropna(inplace=True)
            #min_dist[key]=min_dist[key].T
            min_dist[key]['Distance'] = min_dist[key]['Distance'].astype(float)
            print(f"The keys are {min_dist[key].columns}")
            # Option 1: Keep those within 500m radius
            min_dist[key]=min_dist[key][min_dist[key]['Distance'] <= 500]
            
            # Option 2: For the distance to 2 nearest bus stops
            #min_dist[key] = min_dist[key].nsmallest(2,'Distance')
            print(min_dist[key])
    return min_dist
        
boxes = all_distance(blocks,overall)
print(boxes)
#col = pd.concat(boxes,axis=1)
writer = pd.ExcelWriter('hdb_nearest_bus.xlsx', engine = 'openpyxl',mode='w')
boxes.to_excel(writer, sheet_name = 'Bus Stop')
writer.close()
#boxes = all_distance(blocks,overall)
print(boxes.head(5))


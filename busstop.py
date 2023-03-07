import time
import os
import pandas as pd
import ast
#f = open("./busstopv2.txt", "r")
#pd.read_csv(pd.compat.StringIO("\.join(lines)), sep=";")
with open("busstopv2.txt", "r") as grilled_cheese:
    lines = grilled_cheese.readlines()
    i=0
    collector = []
    for item in lines:
        if i < 1000:
            collector.append(ast.literal_eval(item))
            #df = dict(item)
            #print(df)
            #print(type(df))
            i+=1
    print(len(collector))   
    #dictlines = dict(lines)
    #print(dictlines["value"])

i=2
df = pd.DataFrame(collector[i]['value'],index=[i for i in range(len(collector[1]['value']))])
print(df)
	
import openpyxl
dfs = []
for i in range(len(collector)):
    dfs.append(pd.DataFrame(collector[i]['value'],index=[i for i in range(len(collector[i]['value']))]))
    #print(f"Dataframe {i+1}:\n")
    #print(dfs[i])
print(f"Done with appending {i} DataFrames")
    
busstopdict = pd.concat(dfs,axis=0)
print("Overall DataFrame:")
print(busstopdict.shape)

def director(long,lat):
    reflong = 103.825
    reflat = 1.36076
    #overlap = 0.001
    x = round((long-reflong)*200)
    y = round((lat-reflat)*200)
    return x,y

busstopdict[['UpDown','LeftRight']] = busstopdict.apply(lambda row: director(row['Longitude'],row['Latitude']),axis=1,result_type='expand')

### Now that DataFrame is successfully created, output into Excel spreadsheet:"

# This section is for writing Excel SPREADSHEET (not CSV)
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active

# # For each row in DataFrame
# for r in dataframe_to_rows(busstopdict, index=True, header=True):
#     ws.append(r)

# for cell in ws['A'] + ws[1]:
#     cell.style = 'Pandas'

#wb.save("busstop_overall.xlsx")

### Proceed to compare distance from HDB to each bus stop.
#Load all HDB data first:

# UNBLOCK ON FIRST RUN
tic1 = time.perf_counter()

allhdb = pd.read_excel("./hdb_to_mrt_all.xlsx", index_col=0,engine='openpyxl') #import fresh data  

#global blocks
blocks = allhdb[['postal','lng_hdb','lat_hdb']].drop_duplicates()
blocks[['UpDown','LeftRight']] = blocks.apply(lambda row: director(row['lng_hdb'],row['lat_hdb']),axis=1,result_type='expand')
blocks = blocks[500:2000]

## Small sample trial
#blocks = blocks[(blocks['UpDown']==-7)&(blocks['LeftRight']==0)]
# Find relevant long and lat for HDB (do not mix up with MRT)
print("Blocks DataFrame has shape:",blocks.shape)

#print(blocks)

#blocks = blocks[blocks['postal']=='650383']

def distance(long1,lat1,long2,lat2):
    # 1 deg = 111000m
    dist = 111000*((long2-long1)**2 +(lat2-lat1)**2)**0.5
    return dist
    
def all_distance(hdblist, busstopdict):
    '''Returns closest bus stop from each HDB block in Singapore'''
    all_dist = {}
    min_dist = {}
    alldist_df = pd.DataFrame()

    margin_h = 0.002
    margin_v = 0.002  # of a degree. 1 degree = 111000m
    # 0.004 = 444m region
    box ={}
    # flight - index
    # fdata - block's long and lat

    # For each block in 'blocks'
    global cumtime
    cumtime = 0
    skipped = 0
    calc = 0
    for flight, fdata in hdblist.iterrows():
        #print(f"fdata is:\\n{fdata}")
        #print(f"{fdata.shape}")
        print(f"Working on {fdata['postal']}")
        long1 = fdata['lng_hdb']
        lat1 = fdata['lat_hdb']
        # Compare row-by-row with all bus stops:
        box[flight]=[]
        tic = time.perf_counter()
        for row, busstop in busstopdict.iterrows():
            calc+=1
            if calc%500 ==0:
                print(f'{calc} calculations so far.')
            #print(busstop)
            long2 = busstop['Longitude']
            lat2 = busstop['Latitude']
            #print(f"Long2 is {long2}, Lat2 is {lat2}")
            #print(type(long2),type(lat2))
            if abs(long2-long1) > margin_h:
                skipped +=1
                continue
            elif abs(lat2-lat1) > margin_v:
                skipped +=1
                continue
            elif busstop['UpDown'] != fdata['UpDown']:
                skipped +=1
                continue
            elif busstop['LeftRight'] != fdata['LeftRight']:
                skipped +=1
                continue
            else:
                busstop['Distance']  = distance(long1,lat1,long2,lat2)
                busstop['Postal'] = fdata['postal']
                #print(f"Bus stop {busstop['Description']}: {busstop['Distance']}m")
                box[flight].append(pd.DataFrame(busstop).T)
                #index=[busstop['BusStopCode']
        print(skipped,"skipped rows.\n")
        toc = time.perf_counter()
        cumtime += (toc-tic)
        print("Cumulative time so far:",round(cumtime,2))
        print(f"{skipped} additional rows skipped for {flight}.\n\n")
        
        print("Currently the list of bus stops are:",box[flight])
            #print(f"Distance is {dist}m")

            #dist = math.sqrt(ellipdist(long1, lat1, long, lat)**2)
            #rowdata.insert(loc=0, column='Distance (ft)', value=dist_ft)
        i=0
        for key, value in box.items():
            #print("The value is",value)
            try:
                min_dist[key]= pd.concat(value,axis=0)
            except ValueError:
                continue
            min_dist[key].dropna(inplace=True)
            #min_dist[key]=min_dist[key].T
            min_dist[key]['Distance'] = min_dist[key]['Distance'].astype(float)
            #print(f"The keys are {min_dist[key].columns}")
            # Option 1: Keep those within 500m radius
            min_dist[key]=min_dist[key][min_dist[key]['Distance'] <= 200]
            if i%50 == 0:
                print("Progress so far:",min_dist[key])
            # Option 2: For the distance to 2 nearest bus stops
            #min_dist[key] = min_dist[key].nsmallest(2,'Distance')
            print(min_dist[key])
            i+=1
    return min_dist

# Actual code begins here        
boxes = all_distance(blocks,busstopdict)
print("Cumulative time so far:",round(cumtime,2))

print(boxes)
col = pd.concat(boxes,axis=0)
writer = pd.ExcelWriter('hdb_nearest_bus.xlsx', engine = 'openpyxl',mode='a')
col.to_excel(writer, sheet_name = 'Busstop1000')
writer.close()

toc1 = time.perf_counter()
print(f"{toc1-tic1} seconds taken.")
##boxes = all_distance(blocks,busstopdict)
#print(boxes.head(5))

#busstopdict.to_csv('all_bus_stops.csv',index=True)
